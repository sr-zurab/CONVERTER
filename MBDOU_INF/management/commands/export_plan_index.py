from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from MBDOU_INF.models import PlanPaymentIndex, Organization
import os
from copy import copy
from openpyxl.styles import Border, Side, Alignment


class Command(BaseCommand):
    help = "Экспортирует План ФХД (лист 1) в Excel по организации и году"

    def add_arguments(self, parser):
        parser.add_argument('--org_id', type=int, required=True)
        parser.add_argument('--year', type=int, required=True)

    def handle(self, *args, **options):
        org_id = options['org_id']
        year = options['year']

        try:
            org = Organization.objects.get(id=org_id)
        except Organization.DoesNotExist:
            self.stderr.write(self.style.ERROR(f"Организация с id={org_id} не найдена"))
            return

        base_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.normpath(os.path.join(base_dir, '..', '..', 'templates', 'plan_template.xlsx'))

        if not os.path.exists(template_path):
            self.stderr.write(self.style.ERROR(f"Шаблон не найден по пути: {template_path}"))
            return

        wb = load_workbook(template_path)
        ws = wb['Листы1-5']

        # Сохраняем объединения из строки 31
        header_row = 31
        merged_to_clone = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == header_row and merged_range.max_row == header_row:
                merged_to_clone.append((merged_range.min_col, merged_range.max_col))

        # Получаем координаты 8 нужных столбцов по строке 31
        column_map = {}
        for cell in ws[header_row]:
            if isinstance(cell.value, int) and 1 <= cell.value <= 8:
                column_map[cell.value] = cell.column

        if len(column_map) < 8:
            self.stderr.write("Ошибка: не найдены все 8 столбцов в строке 31 шаблона.")
            return

        columns = [column_map[i] for i in range(1, 9)]

        data = PlanPaymentIndex.objects.filter(organization=org, year=year).order_by('lineCode')

        start_row = 32
        thin = Side(style='thin')

        for i, row in enumerate(data):
            row_index = start_row + i
            values = [
                row.name,
                row.lineCode,
                row.kbk,
                row.analyticCode,
                row.financialYearSum,
                row.planFirstYearSum,
                row.planLastYearSum,
                row.AutPlanYearSumm
            ]
            for j, col in enumerate(columns):
                src_cell = ws.cell(row=header_row, column=col)
                cell = ws.cell(row=row_index, column=col, value=values[j])

                # Копирование стиля и границ
                cell.font = copy(src_cell.font)
                cell.fill = copy(src_cell.fill)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical, wrap_text=True)
                cell.number_format = copy(src_cell.number_format)
                cell.protection = copy(src_cell.protection)
            ws.row_dimensions[row_index].height = 35
            # Копирование объединений
            for min_col, max_col in merged_to_clone:
                col_start = get_column_letter(min_col)
                col_end = get_column_letter(max_col)
                ws.merge_cells(f'{col_start}{row_index}:{col_end}{row_index}')

        # Сохранение в файл
        export_dir = os.path.join(os.path.dirname(template_path), 'exported')
        os.makedirs(export_dir, exist_ok=True)
        export_path = os.path.join(export_dir, f'plan_index_{org_id}_{year}.xlsx')
        wb.save(export_path)

        self.stdout.write(self.style.SUCCESS(f"Экспорт завершён: {export_path}"))