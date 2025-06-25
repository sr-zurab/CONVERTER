# Экспорт плана ФХД в XLSX-файл по шаблону
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from MBDOU_INF.models import PlanPaymentIndex, PlanPaymentTRU, Organization
from copy import copy
import os
import io
import json

# Основная функция экспорта
# org_id — id организации, year — год
# Возвращает файловый объект, имя организации и год

def export_plan_fhd_to_file_object(org_id, year):
    org = Organization.objects.get(id=org_id)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    # Путь к шаблону Excel
    template_path = os.path.normpath(os.path.join(base_dir, '..', 'templates', 'plan_template.xlsx'))
    # Путь к JSON с дефолтными строками
    default_json_path = os.path.normpath(os.path.join(base_dir, '..', '..', 'data', 'default_payment_index.json'))

    wb = load_workbook(template_path)

    # === ЛИСТ 1 ===
    ws1 = wb['Листы1-5']
    header_row_1 = 31  # Номер строки с заголовками
    merged_to_clone_1 = [(r.min_col, r.max_col) for r in ws1.merged_cells.ranges if r.min_row == header_row_1]

    # Сопоставление колонок по заголовкам
    column_map_1 = {
        cell.value: cell.column for cell in ws1[header_row_1] if isinstance(cell.value, int)
    }
    columns_1 = [column_map_1[i] for i in range(1, 9)]
    thin = Side(style='thin')

    # Читаем дефолтные строки из JSON
    with open(default_json_path, encoding='utf-8') as f:
        default_rows = json.load(f)
    # Получаем строки из базы
    db_rows = {str(row.lineCode): row for row in PlanPaymentIndex.objects.filter(organization=org, year=year)}

    # Заполняем строки листа 1
    for i, schema_row in enumerate(default_rows):
        line_code = str(schema_row["lineCode"])
        db_row = db_rows.get(line_code)
        row_index = 32 + i
        values = [
            schema_row["name"],
            schema_row["lineCode"],
            schema_row.get("kbk", ""),
            schema_row.get("analyticCode", ""),
            getattr(db_row, "financialYearSum", "") if db_row else "",
            getattr(db_row, "planFirstYearSum", "") if db_row else "",
            getattr(db_row, "planLastYearSum", "") if db_row else "",
            getattr(db_row, "AutPlanYearSumm", "") if db_row else ""
        ]
        for j, col in enumerate(columns_1):
            src_cell = ws1.cell(row=header_row_1, column=col)
            cell = ws1.cell(row=row_index, column=col, value=values[j])
            cell.font = copy(src_cell.font)
            cell.fill = copy(src_cell.fill)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal,
                                       vertical=src_cell.alignment.vertical, wrap_text=True)
            cell.number_format = copy(src_cell.number_format)
            cell.protection = copy(src_cell.protection)
        ws1.row_dimensions[row_index].height = 35
        for min_col, max_col in merged_to_clone_1:
            ws1.merge_cells(f'{get_column_letter(min_col)}{row_index}:{get_column_letter(max_col)}{row_index}')

    # === ЛИСТ 2 ===
    ws2 = wb['Листы6-7']
    header_row_2 = 8  # Номер строки с заголовками
    column_letters_2 = ['A', 'F', 'BD', 'BJ', 'BP', 'BX', 'CF', 'CN']
    columns_2 = [column_index_from_string(letter) for letter in column_letters_2]
    merged_to_clone_2 = [(r.min_col, r.max_col) for r in ws2.merged_cells.ranges if r.min_row == header_row_2]

    # Получаем строки из базы для листа 2
    data_2 = PlanPaymentTRU.objects.filter(organization=org, year=year).order_by('lineCode')

    # Заполняем строки листа 2
    for i, row in enumerate(data_2):
        row_index = 9 + i
        values = [
            row.lineNum,
            row.name,
            row.lineCode,
            row.yearStart,
            row.financialYearSum,
            row.planFirstYearSum,
            row.planLastYearSum,
            row.AutPlanYearSumm
        ]
        for j, col in enumerate(columns_2):
            src_cell = ws2.cell(row=header_row_2, column=col)
            cell = ws2.cell(row=row_index, column=col, value=values[j])
            cell.font = copy(src_cell.font)
            cell.fill = copy(src_cell.fill)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal,
                                       vertical=src_cell.alignment.vertical, wrap_text=True)
            cell.number_format = copy(src_cell.number_format)
            cell.protection = copy(src_cell.protection)
        ws2.row_dimensions[row_index].height = 35
        for min_col, max_col in merged_to_clone_2:
            ws2.merge_cells(f'{get_column_letter(min_col)}{row_index}:{get_column_letter(max_col)}{row_index}')

    # Сохраняем рабочую книгу во временный файловый объект
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream, org.name.replace(" ", "_"), year  # Возвращаем файл, имя организации и год