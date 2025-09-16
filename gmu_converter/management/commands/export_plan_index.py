from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from gmu_converter.models import PlanPaymentIndex, PlanPaymentTRU, Organization
import os
from copy import copy
from openpyxl.styles import Border, Side, Alignment


class Command(BaseCommand):
    help = "Экспортирует План ФХД (листы 1 и 2) в Excel по организации и году"

    def add_arguments(self, parser):
        parser.add_argument('--org_id', type=int, required=True)
        parser.add_argument('--year', type=int, required=True)

    def handle(self, *args, **options):
        org_id = options['org_id']
        year = options['year']

        try:
            org = Organization.objects.get(id=org_id)
        except Organization.DoesNotExist:
            self.stderr.write(self.style.ERROR(f"Организация с id={org_id} не найдена"))
            return

        base_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.normpath(os.path.join(base_dir, '..', '..', 'templates', 'plan_template.xlsx'))

        if not os.path.exists(template_path):
            self.stderr.write(self.style.ERROR(f"Шаблон не найден по пути: {template_path}"))
            return

        wb = load_workbook(template_path)

        # === ЛИСТ 1 ===
        ws1 = wb['Листы1-5']
        header_row_1 = 31
        merged_to_clone_1 = []
        for merged_range in ws1.merged_cells.ranges:
            if merged_range.min_row == header_row_1 and merged_range.max_row == header_row_1:
                merged_to_clone_1.append((merged_range.min_col, merged_range.max_col))

        column_map_1 = {}
        for cell in ws1[header_row_1]:
            if isinstance(cell.value, int) and 1 <= cell.value <= 8:
                column_map_1[cell.value] = cell.column

        if len(column_map_1) < 8:
            self.stderr.write("Ошибка: не найдены все 8 столбцов в строке 31 шаблона.")
            return

        columns_1 = [column_map_1[i] for i in range(1, 9)]
        data_1 = PlanPaymentIndex.objects.filter(organization=org, year=year).order_by('lineCode')

        start_row_1 = 32
        thin = Side(style='thin')

        for i, row in enumerate(data_1):
            row_index = start_row_1 + i
            values = [
                row.name,
                row.lineCode,
                row.kbk,
                row.analyticCode,
                row.financialYearSum,
                row.planFirstYearSum,
                row.planLastYearSum,
                row.AutPlanYearSumm
            ]
            for j, col in enumerate(columns_1):
                src_cell = ws1.cell(row=header_row_1, column=col)
                cell = ws1.cell(row=row_index, column=col, value=values[j])
                cell.font = copy(src_cell.font)
                cell.fill = copy(src_cell.fill)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical, wrap_text=True)
                cell.number_format = copy(src_cell.number_format)
                cell.protection = copy(src_cell.protection)
            ws1.row_dimensions[row_index].height = 35
            for min_col, max_col in merged_to_clone_1:
                col_start = get_column_letter(min_col)
                col_end = get_column_letter(max_col)
                ws1.merge_cells(f'{col_start}{row_index}:{col_end}{row_index}')

        # === ЛИСТ 2 ===
        ws2 = wb['Листы6-7']
        header_row_2 = 8
        start_row_2 = 9

        # Ячейки заголовков: A8, F8, BD8, BJ8, BP8, BX8, CF8, CN8
        column_letters_2 = ['A', 'F', 'BD', 'BJ', 'BP', 'BX', 'CF', 'CN']
        columns_2 = [column_index_from_string(letter) for letter in column_letters_2]

        column_map_2 = {}
        for col_num, cell in zip(range(1, 9), columns_2):
            column_map_2[col_num] = cell

        merged_to_clone_2 = []
        for merged_range in ws2.merged_cells.ranges:
            if merged_range.min_row == header_row_2 and merged_range.max_row == header_row_2:
                merged_to_clone_2.append((merged_range.min_col, merged_range.max_col))

        data_2 = PlanPaymentTRU.objects.filter(organization=org, year=year).order_by('lineCode')

        for i, row in enumerate(data_2):
            row_index = start_row_2 + i
            values = [
                row.lineNum,
                row.name,
                row.lineCode,
                row.yearStart,
                row.financialYearSum,
                row.planFirstYearSum,
                row.planLastYearSum,
                row.AutPlanYearSumm
            ]
            for j, col in enumerate(columns_2):
                src_cell = ws2.cell(row=header_row_2, column=col)
                cell = ws2.cell(row=row_index, column=col, value=values[j])
                cell.font = copy(src_cell.font)
                cell.fill = copy(src_cell.fill)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical, wrap_text=True)
                cell.number_format = copy(src_cell.number_format)
                cell.protection = copy(src_cell.protection)
            ws2.row_dimensions[row_index].height = 35
            for min_col, max_col in merged_to_clone_2:
                col_start = get_column_letter(min_col)
                col_end = get_column_letter(max_col)
                ws2.merge_cells(f'{col_start}{row_index}:{col_end}{row_index}')

        # === Сохранение ===
        export_dir = os.path.join(os.path.dirname(template_path), 'exported')
        os.makedirs(export_dir, exist_ok=True)
        export_path = os.path.join(export_dir, f'plan_index_{org_id}_{year}.xlsx')
        wb.save(export_path)

        self.stdout.write(self.style.SUCCESS(f"Экспорт завершён: {export_path}"))